import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import json
import urllib.request

# ============================================================
# LICENCIA REMOTA
# ============================================================

LICENSE_URL = "https://api.github.com/gists/0550fbbc1474f8110bb0555629eb2362"


def verificar_licencia():
    try:
        req = urllib.request.Request(LICENSE_URL, headers={
            'Accept': 'application/vnd.github.v3+json',
            'Cache-Control': 'no-cache',
        })
        with urllib.request.urlopen(req, timeout=8) as resp:
            gist = json.loads(resp.read().decode('utf-8'))
            content = list(gist['files'].values())[0]['content']
            data = json.loads(content)
            return data.get('activo', True), data.get('mensaje', '')
    except Exception:
        return True, ''


# ============================================================
# CONFIGURACIÓN
# ============================================================

MESES_ES = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
    5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
    9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE',
}

EXAM_NAMES = {
    'GRUPO_SANGUINEO': 'GRUPO SANGUÍNEO Y FACTOR RH',
    'SOMNOLENCIA': 'TEST DE SOMNOLENCIA Y FATIGA - EPWORTH',
    'ESPIROMETRIA': 'ESPIROMETRÍA',
    'YOSHITAKE': 'TEST DE YOSHITAKE',
    'PSICOSENSOMETRICO': 'PSICOSENSOMETRICO',
}

PRICE_FORMAT = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
DATE_FORMAT = 'mm-dd-yy'

FONT_DATA = Font(name='Calibri', size=9)
FONT_TITLE = Font(name='Calibri', size=9, bold=True)
FONT_HEADER = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

COLUMN_WIDTHS = {
    1: 5.0, 2: 10.5, 3: 7.5, 4: 16.6, 5: 29.5, 6: 13.0, 7: 5.5,
    8: 11.5, 9: 17.6, 10: 16.3, 11: 15.6, 12: 10.0, 13: 10.0,
    14: 10.7, 15: 10.9, 16: 10.9, 17: 102.5,
}

HEADERS_OUT = [
    'N°', 'Fecha', 'DPTO.', 'Centro médico', 'NOMBRE COMPLETO',
    'Nro. Doc.', 'Edad', 'Sexo', 'Area de trabajo', 'Puesto de trabajo',
    'Unidad de Negocio', 'Tipo de evaluación', 'Centro de Costos',
    'Proyecto/sede', 'Perfil', ' PRECIO ', 'Observacion',
]

OCUP_COLUMN_RULES = {
    'FECHA':             lambda h: h == 'FECHA',
    'SEDE':              lambda h: h == 'SEDE',
    'PACIENTE':          lambda h: h == 'PACIENTE',
    'DNI':               lambda h: h == 'DNI',
    'EDAD':              lambda h: h == 'EDAD',
    'SEXO':              lambda h: h == 'SEXO',
    'AREA':              lambda h: h in ('AREA', 'ÁREA'),
    'PUESTO':            lambda h: h == 'PUESTO',
    'LUGAR_TRABAJO':     lambda h: 'LUGAR DE TRABAJO' in h,
    'UNIDAD_NEGOCIO':    lambda h: 'UNIDAD DE NEGOCIO' in h,
    'CENTRO_COSTOS':     lambda h: 'CENTRO DE COSTO' in h,
    'PROYECTO':          lambda h: h == 'PROYECTO',
    'TIPO_EXAMEN':       lambda h: 'TIPO DE EXAMEN' in h,
    'PERFIL':            lambda h: h == 'PERFIL',
    'PRECIO':            lambda h: h == 'PRECIO',
    'GRUPO_SANGUINEO':   lambda h: 'GRUPO SANGU' in h,
    'SOMNOLENCIA':       lambda h: 'SOMNOLENCIA' in h,
    'ESPIROMETRIA':      lambda h: 'ESPIROMETR' in h,
    'PSICOSENSOMETRICO': lambda h: h == 'PSICOSENSOMETRICO',
    'YOSHITAKE':         lambda h: 'YOSHITAKE' in h,
}

ASIS_COLUMN_RULES = {
    'FECHA':          lambda h: h == 'FECHA',
    'SEDE':           lambda h: h == 'SEDE',
    'PROYECTO':       lambda h: h == 'PROYECTO',
    'AREA':           lambda h: h in ('AREA', 'ÁREA'),
    'PUESTO':         lambda h: h == 'PUESTO',
    'PACIENTE':       lambda h: h == 'PACIENTE',
    'DNI':            lambda h: h == 'DNI',
    'EDAD':           lambda h: h == 'EDAD',
    'SEXO':           lambda h: h == 'SEXO',
    'PRECIO':         lambda h: h == 'PRECIO',
    'PANEL_5D':       lambda h: 'PANEL 5D' in h or 'DROGAS' in h,
    'COCAINA':        lambda h: 'COCAINA' in h,
    'MARIHUANA':      lambda h: 'MARIHUANA' in h,
    'METANFETAMINA':  lambda h: 'METANFETAMINA' in h,
}


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def safe_str(val, default=''):
    if val is None:
        return default
    if isinstance(val, float) and pd.isna(val):
        return default
    s = str(val).strip()
    return default if s.lower() == 'nan' else s


def safe_float(val, default=0.0):
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return default


def convertir_sexo(sexo):
    s = str(sexo).strip().upper()
    if s == 'M':
        return 'MASCULINO'
    if s == 'F':
        return 'FEMENINO'
    return sexo


def convertir_fecha(fecha_val):
    if isinstance(fecha_val, datetime):
        return fecha_val
    try:
        return datetime.strptime(str(fecha_val).strip(), '%d-%m-%Y')
    except (ValueError, TypeError):
        try:
            return pd.to_datetime(fecha_val)
        except Exception:
            return None


def obtener_perfil_numero(perfil_texto):
    texto = str(perfil_texto).upper().replace('PERFIL', '').strip()
    try:
        return int(texto.split()[0])
    except (ValueError, IndexError):
        return None


def tiene_examen(row, col_map, key):
    if key not in col_map:
        return False
    val = row[col_map[key]]
    return str(val).strip() in ('1', '1.0')


def normalizar_dni(dni):
    return str(dni).strip().replace('.0', '').lstrip('0')


def safe_val(row, col_map, key, as_str=True):
    if key not in col_map:
        return ''
    val = row[col_map[key]]
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    if as_str:
        s = str(val).strip()
        return '' if s.lower() == 'nan' else s
    return val


# ============================================================
# LECTURA DE ARCHIVOS
# ============================================================

def encontrar_fila_encabezados(df, keyword='PACIENTE'):
    for i in range(min(15, len(df))):
        vals = [str(v).strip().upper() for v in df.iloc[i]]
        if any(keyword in v for v in vals):
            return i
    return None


def detectar_columnas(df, header_row, rules):
    headers = [str(v).strip().upper() for v in df.iloc[header_row]]
    col_map = {}
    for key, match_fn in rules.items():
        for idx, h in enumerate(headers):
            if match_fn(h):
                col_map[key] = idx
                break
    missing = [k for k in rules if k not in col_map]
    return col_map, missing


def encontrar_hoja_datos(content):
    """Busca la hoja con datos crudos (no tabla dinámica)."""
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        best_sheet = None
        best_cols = 0
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.max_column <= best_cols:
                continue
            for r in range(1, min(10, ws.max_row + 1)):
                found = False
                for c in ws[r]:
                    if c.value and 'PACIENTE' in str(c.value).upper():
                        best_cols = ws.max_column
                        best_sheet = name
                        found = True
                        break
                if found:
                    break
        return best_sheet
    except Exception:
        return None


def leer_excel_flexible(content, sheet_name=None):
    readers = []
    if sheet_name:
        readers.append(lambda: pd.read_excel(
            io.BytesIO(content), sheet_name=sheet_name,
            header=None, engine='openpyxl'))
    readers.extend([
        lambda: pd.read_excel(io.BytesIO(content), header=None, engine='openpyxl'),
        lambda: pd.read_excel(io.BytesIO(content), header=None, engine='xlrd'),
        lambda: pd.read_html(io.BytesIO(content), encoding='utf-8', flavor='lxml')[0],
        lambda: pd.read_html(io.BytesIO(content), flavor='html5lib')[0],
    ])
    for reader in readers:
        try:
            df = reader()
            df.columns = range(df.shape[1])
            return df
        except Exception:
            continue
    raise RuntimeError("No se pudo leer el archivo.")


def extraer_datos(df, rules):
    header_row = encontrar_fila_encabezados(df)
    if header_row is None:
        raise RuntimeError("No se encontró fila de encabezados (PACIENTE).")
    col_map, missing = detectar_columnas(df, header_row, rules)

    data = df.iloc[header_row + 1:].copy()
    data.columns = range(df.shape[1])
    data = data[data[0].astype(str).str.strip().str.upper() != 'TOTAL']
    data = data[pd.to_numeric(data[0], errors='coerce').notna()]
    data.reset_index(drop=True, inplace=True)
    return data, col_map, missing


def leer_ocupacional(uploaded_file):
    content = uploaded_file.read()
    uploaded_file.seek(0)
    sheet_name = encontrar_hoja_datos(content)
    df = leer_excel_flexible(content, sheet_name)
    return extraer_datos(df, OCUP_COLUMN_RULES)


def leer_asistencial(uploaded_file):
    content = uploaded_file.read()
    uploaded_file.seek(0)
    sheet_name = encontrar_hoja_datos(content)
    df = leer_excel_flexible(content, sheet_name)
    return extraer_datos(df, ASIS_COLUMN_RULES)


def clasificar_archivos(files):
    """Auto-detecta cuál archivo es Ocupacional y cuál Asistencial."""
    ocup = asis = None

    for f in files:
        name = f.name.upper()
        if 'ASISTENCIAL' in name:
            asis = f
        elif 'OCUPACIONAL' in name:
            ocup = f

    if ocup and asis:
        return ocup, asis

    remaining = [f for f in files if f is not ocup and f is not asis]
    for f in remaining:
        content = f.read()
        f.seek(0)
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
            all_headers = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for r in range(1, min(10, ws.max_row + 1)):
                    for c in ws[r]:
                        if c.value:
                            all_headers.append(str(c.value).upper())

            has_drugs = any(
                'DROGA' in h or 'COCAINA' in h or 'MARIHUANA' in h
                for h in all_headers
            )
            has_ocup = any(
                'PSICOSENSOMETRICO' in h or 'UNIDAD DE NEGOCIO' in h
                for h in all_headers
            )

            if has_drugs and not asis:
                asis = f
            elif has_ocup and not ocup:
                ocup = f
            elif not ocup:
                ocup = f
            elif not asis:
                asis = f
        except Exception:
            if not ocup:
                ocup = f
            elif not asis:
                asis = f

    return ocup, asis


# ============================================================
# PROCESAMIENTO
# ============================================================

def procesar_datos(data_ocup, col_ocup, data_asis, col_asis,
                   precios, precios_toxi):
    dni_lookup = {}
    for _, row in data_ocup.iterrows():
        dni = normalizar_dni(safe_val(row, col_ocup, 'DNI'))
        if dni:
            dni_lookup[dni] = row

    periodicos = []
    retiro = []
    adicionales_reg = []

    for _, row in data_ocup.iterrows():
        perfil_texto = safe_val(row, col_ocup, 'PERFIL')
        perfil_num = obtener_perfil_numero(perfil_texto)
        tipo_examen = safe_val(row, col_ocup, 'TIPO_EXAMEN').upper()
        precio_total = safe_float(safe_val(row, col_ocup, 'PRECIO'))

        has_grupo = tiene_examen(row, col_ocup, 'GRUPO_SANGUINEO')
        has_somn = tiene_examen(row, col_ocup, 'SOMNOLENCIA')
        has_yosh = tiene_examen(row, col_ocup, 'YOSHITAKE')
        has_psico = tiene_examen(row, col_ocup, 'PSICOSENSOMETRICO')

        precio_adicional = 0
        obs_adic_parts = []
        if perfil_num is not None and perfil_num <= 4:
            if has_grupo:
                precio_adicional += precios['GRUPO_SANGUINEO']
                obs_adic_parts.append(EXAM_NAMES['GRUPO_SANGUINEO'])
            if has_somn:
                precio_adicional += precios['SOMNOLENCIA']
                obs_adic_parts.append(EXAM_NAMES['SOMNOLENCIA'])
            if has_yosh:
                precio_adicional += precios['YOSHITAKE']
                obs_adic_parts.append(EXAM_NAMES['YOSHITAKE'])
            if has_psico:
                precio_adicional += precios['PSICOSENSOMETRICO']
                obs_adic_parts.append(EXAM_NAMES['PSICOSENSOMETRICO'])

        obs_emo_parts = []
        if has_grupo:
            obs_emo_parts.append(EXAM_NAMES['GRUPO_SANGUINEO'])
        if has_somn:
            obs_emo_parts.append(EXAM_NAMES['SOMNOLENCIA'])
        if has_yosh:
            obs_emo_parts.append(EXAM_NAMES['YOSHITAKE'])
        if has_psico:
            obs_emo_parts.append(EXAM_NAMES['PSICOSENSOMETRICO'])

        unidad = safe_val(row, col_ocup, 'UNIDAD_NEGOCIO')
        cc_raw = safe_val(row, col_ocup, 'CENTRO_COSTOS', as_str=False)
        centro_costos = cc_raw if cc_raw != '' else ''
        sede = safe_val(row, col_ocup, 'SEDE')

        paciente = {
            'fecha': convertir_fecha(safe_val(row, col_ocup, 'FECHA', as_str=False)),
            'dpto': 'LIMA',
            'centro_medico': sede,
            'nombre': safe_val(row, col_ocup, 'PACIENTE'),
            'dni': safe_val(row, col_ocup, 'DNI', as_str=False),
            'edad': safe_val(row, col_ocup, 'EDAD', as_str=False),
            'sexo': convertir_sexo(safe_val(row, col_ocup, 'SEXO')),
            'area': safe_val(row, col_ocup, 'AREA'),
            'puesto': safe_val(row, col_ocup, 'PUESTO'),
            'unidad_negocio': unidad,
            'tipo_evaluacion': tipo_examen,
            'centro_costos': centro_costos,
            'proyecto_sede': safe_val(row, col_ocup, 'PROYECTO'),
            'perfil': perfil_num,
            'precio_total': precio_total,
            'precio_base': precio_total - precio_adicional,
            'precio_adicional': precio_adicional,
            'obs_emo': ' + '.join(obs_emo_parts),
            'obs_adic': ' + '.join(obs_adic_parts),
        }

        if tipo_examen == 'RETIRO':
            retiro.append(paciente)
        else:
            periodicos.append(paciente)

        if precio_adicional > 0 and tipo_examen != 'RETIRO':
            adicionales_reg.append(paciente)

    adicionales_toxi = []

    if data_asis is not None and col_asis is not None:
        for _, row in data_asis.iterrows():
            has_panel5d = tiene_examen(row, col_asis, 'PANEL_5D')
            has_coca = tiene_examen(row, col_asis, 'COCAINA')
            has_mari = tiene_examen(row, col_asis, 'MARIHUANA')
            has_meta = tiene_examen(row, col_asis, 'METANFETAMINA')

            if not (has_panel5d or has_coca or has_mari or has_meta):
                continue

            precio_preliq = safe_float(safe_val(row, col_asis, 'PRECIO'))

            precio_toxi_adic = (precios_toxi['PANEL_5D']
                                if has_panel5d
                                else precios_toxi['COCA_MARI_META'])

            dni_asis = normalizar_dni(safe_val(row, col_asis, 'DNI'))
            ocup_row = dni_lookup.get(dni_asis)

            area = safe_val(row, col_asis, 'AREA')
            puesto = safe_val(row, col_asis, 'PUESTO')
            sede = safe_val(row, col_asis, 'SEDE')
            proyecto = safe_val(row, col_asis, 'PROYECTO')
            unidad = ''
            centro_costos = ''
            perfil_toxi = None

            if ocup_row is not None:
                if not area:
                    area = safe_val(ocup_row, col_ocup, 'AREA')
                if not puesto:
                    puesto = safe_val(ocup_row, col_ocup, 'PUESTO')
                if not sede:
                    sede = safe_val(ocup_row, col_ocup, 'SEDE')
                if not proyecto or proyecto.lower() == 'sin proyecto':
                    v = safe_val(ocup_row, col_ocup, 'PROYECTO')
                    if v:
                        proyecto = v
                unidad = safe_val(ocup_row, col_ocup, 'UNIDAD_NEGOCIO')
                cc_raw = safe_val(ocup_row, col_ocup, 'CENTRO_COSTOS', as_str=False)
                centro_costos = cc_raw if cc_raw != '' else ''
                pf = safe_val(ocup_row, col_ocup, 'PERFIL')
                if pf:
                    perfil_toxi = obtener_perfil_numero(pf)

            pac_toxi = {
                'fecha': convertir_fecha(
                    safe_val(row, col_asis, 'FECHA', as_str=False)),
                'dpto': 'LIMA',
                'centro_medico': sede,
                'nombre': safe_val(row, col_asis, 'PACIENTE'),
                'dni': safe_val(row, col_asis, 'DNI', as_str=False),
                'edad': safe_val(row, col_asis, 'EDAD', as_str=False),
                'sexo': convertir_sexo(safe_val(row, col_asis, 'SEXO')),
                'area': area,
                'puesto': puesto,
                'unidad_negocio': unidad,
                'centro_costos': centro_costos,
                'proyecto_sede': proyecto,
                'perfil_toxi': perfil_toxi,
                'precio_toxi': precio_preliq,
                'precio_adic_toxi': precio_toxi_adic,
            }

            adicionales_toxi.append(pac_toxi)

    def sort_key(p):
        fecha = p.get('fecha') or datetime(2099, 1, 1)
        return (fecha, p.get('nombre', ''))

    periodicos.sort(key=sort_key)
    retiro.sort(key=sort_key)
    adicionales_reg.sort(key=sort_key)
    adicionales_toxi.sort(key=sort_key)

    return periodicos, retiro, adicionales_reg, adicionales_toxi


# ============================================================
# GENERACIÓN DEL EXCEL
# ============================================================

def setup_hoja(ws):
    c = ws.cell(row=4, column=4, value='MOTORES DIESEL ANDINOS S.A.- MODASA')
    c.font = FONT_TITLE
    for col_idx, header in enumerate(HEADERS_OUT, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def escribir_fila(ws, row_num, valores):
    for col_idx, valor in enumerate(valores, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=valor)
        cell.font = FONT_DATA
        if col_idx == 2 and isinstance(valor, datetime):
            cell.number_format = DATE_FORMAT
        elif col_idx == 16:
            cell.number_format = PRICE_FORMAT


def escribir_datos(ws, pacientes, campo_precio, campo_obs, start_row=7):
    for i, pac in enumerate(pacientes, 1):
        row_num = start_row + i - 1
        obs = pac.get(campo_obs, '') if campo_obs else ''
        valores = [
            i, pac.get('fecha'), pac.get('dpto', 'LIMA'),
            pac.get('centro_medico', ''), pac.get('nombre', ''),
            pac.get('dni', ''), pac.get('edad', ''), pac.get('sexo', ''),
            pac.get('area', ''), pac.get('puesto', ''),
            pac.get('unidad_negocio', ''), pac.get('tipo_evaluacion', ''),
            pac.get('centro_costos', ''), pac.get('proyecto_sede', ''),
            pac.get('perfil'), pac.get(campo_precio, 0), obs,
        ]
        escribir_fila(ws, row_num, valores)
    return start_row + len(pacientes) - 1 if pacientes else start_row - 1


def escribir_totales(ws, first_data_row, last_data_row):
    col = 'P'
    if last_data_row < first_data_row:
        s = first_data_row + 1
        for label, row_offset in [('B. INPONIBLE', 0), ('IGV', 1), ('TOTAL', 2)]:
            ws.cell(row=s + row_offset, column=15, value=label).font = FONT_DATA
            c = ws.cell(row=s + row_offset, column=16, value=0)
            c.font = FONT_DATA
            c.number_format = PRICE_FORMAT
        return

    s = last_data_row + 2
    ws.cell(row=s, column=15, value='B. IMPONIBLE').font = FONT_DATA
    c = ws.cell(row=s, column=16)
    c.value = f'=SUM({col}{first_data_row}:{col}{last_data_row})'
    c.font = FONT_DATA
    c.number_format = PRICE_FORMAT

    ws.cell(row=s + 2, column=15, value='TOTAL').font = FONT_DATA
    c = ws.cell(row=s + 2, column=16)
    c.value = f'=+{col}{s}*1.18'
    c.font = FONT_DATA
    c.number_format = PRICE_FORMAT

    ws.cell(row=s + 1, column=15, value='IGV').font = FONT_DATA
    c = ws.cell(row=s + 1, column=16)
    c.value = f'=+{col}{s + 2}-{col}{s}'
    c.font = FONT_DATA
    c.number_format = PRICE_FORMAT


def generar_excel(periodicos, retiro, adicionales_reg, adicionales_toxi):
    wb = Workbook()
    wb.remove(wb.active)

    # EMO DOKTUZ — periódicos, precio base (sin adicionales)
    ws = wb.create_sheet(title='EMO DOKTUZ')
    setup_hoja(ws)
    lr = escribir_datos(ws, periodicos, 'precio_base', 'obs_emo')
    escribir_totales(ws, 7, lr)

    # EMOR DOKTUZ — retiro
    wsr = wb.create_sheet(title='EMOR DOKTUZ')
    setup_hoja(wsr)
    lrr = escribir_datos(wsr, retiro, 'precio_total', None)
    escribir_totales(wsr, 7, lrr)

    # ADICIONALES DOKTUZ — regulares + toxicológicos al final
    wsa = wb.create_sheet(title='ADICIONALES DOKTUZ')
    setup_hoja(wsa)
    lra = escribir_datos(wsa, adicionales_reg, 'precio_adicional', 'obs_adic')

    toxi_start = (lra + 1) if adicionales_reg else 7
    n_start = len(adicionales_reg) + 1
    for i, pac in enumerate(adicionales_toxi):
        row_num = toxi_start + i
        vals = [
            n_start + i, pac.get('fecha'), pac.get('dpto', 'LIMA'),
            pac.get('centro_medico', ''), pac.get('nombre', ''),
            pac.get('dni', ''), pac.get('edad', ''), pac.get('sexo', ''),
            pac.get('area', ''), pac.get('puesto', ''),
            pac.get('unidad_negocio', ''), 'ADICIONALES',
            pac.get('centro_costos', ''), pac.get('proyecto_sede', ''),
            'LABORATORIO', pac.get('precio_adic_toxi', 0), 'TOXICOLOGICO',
        ]
        escribir_fila(wsa, row_num, vals)

    final_row = toxi_start + len(adicionales_toxi) - 1 if adicionales_toxi else lra
    if not adicionales_reg and not adicionales_toxi:
        final_row = 6
    escribir_totales(wsa, 7, final_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# INTERFAZ WEB (STREAMLIT)
# ============================================================

st.set_page_config(page_title="Valorizado MODASA", layout="centered")

activo, mensaje_bloqueo = verificar_licencia()
if not activo:
    st.error(mensaje_bloqueo or "Aplicación desactivada. Contacte al proveedor.")
    st.stop()

st.title("Valorizado MODASA")
st.caption("Transforma Preliquidación en Valorizado formato MODASA (.xlsx)")
st.divider()

uploaded_files = st.file_uploader(
    "Sube los archivos de Preliquidación (Ocupacional y opcionalmente Asistencial)",
    type=['xls', 'xlsx'],
    accept_multiple_files=True,
    help="Sube 1 archivo (Ocupacional) o 2 archivos (Ocupacional + Asistencial)",
)

st.subheader("Precios de Adicionales (S/)")
c1, c2 = st.columns(2)
with c1:
    p_grupo = st.number_input("Grupo Sanguíneo y Factor RH", value=5.0,
                              min_value=0.0, step=1.0)
    p_somn = st.number_input("Test de Somnolencia (Epworth)", value=3.0,
                             min_value=0.0, step=1.0)
with c2:
    p_psico = st.number_input("Psicosensométrico", value=25.0,
                              min_value=0.0, step=1.0)
    p_yosh = st.number_input("Test de Yoshitake", value=2.0,
                             min_value=0.0, step=1.0)

precios = {
    'GRUPO_SANGUINEO': p_grupo,
    'PSICOSENSOMETRICO': p_psico,
    'SOMNOLENCIA': p_somn,
    'YOSHITAKE': p_yosh,
}

st.subheader("Precios de Toxicológico (S/)")
c3, c4 = st.columns(2)
with c3:
    p_panel5d = st.number_input("Panel 5D", value=32.0,
                                min_value=0.0, step=1.0)
with c4:
    p_paquete = st.number_input("Cocaína + Marihuana + Metanfetamina",
                                value=32.0, min_value=0.0, step=1.0)

precios_toxi = {
    'PANEL_5D': p_panel5d,
    'COCA_MARI_META': p_paquete,
}

st.subheader("Configuración del archivo")
c5, c6, c7 = st.columns(3)
with c5:
    mes_idx = st.selectbox("Mes", options=list(MESES_ES.keys()),
                           format_func=lambda x: MESES_ES[x],
                           index=datetime.now().month - 1)
with c6:
    anio = st.number_input("Año", value=datetime.now().year,
                           min_value=2020, max_value=2035)
with c7:
    version = st.text_input("Versión", value="V3")

st.divider()

if uploaded_files and len(uploaded_files) >= 1:
    if st.button("Generar Valorizado", type="primary", use_container_width=True):
        with st.spinner("Procesando…"):
            try:
                ocup_file = None
                asis_file = None
                data_asis = None
                col_asis = None

                if len(uploaded_files) >= 2:
                    ocup_file, asis_file = clasificar_archivos(uploaded_files)
                    if not ocup_file or not asis_file:
                        st.error("No se pudo identificar cuál archivo es "
                                 "Ocupacional y cuál Asistencial. "
                                 "Verifica los nombres de los archivos.")
                        st.stop()
                    st.info(f"**Ocupacional:** {ocup_file.name}  \n"
                            f"**Asistencial:** {asis_file.name}")
                else:
                    ocup_file = uploaded_files[0]
                    st.info(f"**Ocupacional:** {ocup_file.name}  \n"
                            f"**Asistencial:** no proporcionado (sin toxicológicos)")

                data_ocup, col_ocup, miss_o = leer_ocupacional(ocup_file)

                if asis_file:
                    data_asis, col_asis, miss_a = leer_asistencial(asis_file)
                    if miss_a:
                        st.warning(f"Columnas no encontradas (Asistencial): "
                                   f"{miss_a}")

                if miss_o:
                    st.warning(f"Columnas no encontradas (Ocupacional): "
                               f"{miss_o}")

                (periodicos, retiro,
                 adicionales_reg,
                 adicionales_toxi) = procesar_datos(
                    data_ocup, col_ocup, data_asis, col_asis,
                    precios, precios_toxi)

                total_emo = sum(p['precio_base'] for p in periodicos)
                total_emor = sum(p['precio_total'] for p in retiro)
                total_adic_r = sum(p['precio_adicional']
                                   for p in adicionales_reg)
                total_adic_t = sum(p['precio_adic_toxi']
                                   for p in adicionales_toxi)
                total_adic = total_adic_r + total_adic_t

                excel_bytes = generar_excel(
                    periodicos, retiro,
                    adicionales_reg, adicionales_toxi)

                st.success("Valorizado generado exitosamente.")

                cols = st.columns(3)
                cols[0].metric("EMO DOKTUZ",
                               f"{len(periodicos)} pac.",
                               f"S/ {total_emo:,.2f}")
                cols[1].metric("EMOR DOKTUZ",
                               f"{len(retiro)} pac.",
                               f"S/ {total_emor:,.2f}")
                cols[2].metric("ADICIONALES",
                               f"{len(adicionales_reg)+len(adicionales_toxi)}"
                               f" pac.",
                               f"S/ {total_adic:,.2f}")

                st.metric("TOTAL FACTURACIÓN",
                          f"S/ {total_emo + total_emor + total_adic:,.2f}")

                file_name = (f"Valorizado Doktuz - Modasa - "
                             f"{MESES_ES[mes_idx]} {anio} {version}.xlsx")
                st.download_button(
                    label=f"Descargar {file_name}",
                    data=excel_bytes,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument"
                         ".spreadsheetml.sheet",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Error: {e}")

else:
    st.info("Sube al menos el archivo de Preliquidación Ocupacional para comenzar.")
